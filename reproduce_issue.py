
import json
import sys
import os
from openpyxl import load_workbook

# Add backend to path
sys.path.append(os.path.join(os.getcwd(), "backend"))

from utils.json_to_excel import dynamic_json_to_excel

def test_reproduction():
    # Simulate the data exactly as in dscr_template_processor.py

    # 1. Create a simulated item
    item = {
        "rule_id": "Purchase",
        "dscr_parameters": "Purchase",
        "category": "Eligible Transactions",
        "sub_category": "Feature Eligibility",
        "ppe_field_type": "Hard",
        "guideline_1": "Allowed",
        "guideline_2": "Allowed",
        "comparison_notes": "Notes"
    }

    # 2. Create results list
    results = [item]

    # 3. Add s_no
    for idx, it in enumerate(results, 1):
        it["s_no"] = idx

    # 4. Define params
    header_map = {
        "s_no": "S.No",
        "dscr_parameters": "DSCR PARAMETERS",
        "category": "Category",
        "sub_category": "Sub Category",
        "guideline_1": "Guideline 1",
        "guideline_2": "Guideline 2",
        "comparison_notes": "Comparison Notes"
    }

    column_order = [
        "s_no",
        "dscr_parameters",
        "category",
        "sub_category",
        "guideline_1",
        "guideline_2",
        "comparison_notes"
    ]

    hidden_columns = ["rule_id", "ppe_field_type"]

    output_path = "reproduce_output.xlsx"
    if os.path.exists(output_path):
        os.remove(output_path)

    # 5. Call function
    dynamic_json_to_excel(
        results,
        output_path,
        header_map=header_map,
        hidden_columns=hidden_columns,
        column_order=column_order
    )

    # 6. Check output
    wb = load_workbook(output_path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    print(f"Headers: {headers}")

    # Check for duplicates
    seen = set()
    dupes = []
    for h in headers:
        if h in seen:
            dupes.append(h)
        seen.add(h)

    if dupes:
        print(f"❌ FOUND DUPLICATES: {dupes}")
    else:
        print("✅ No duplicates found.")

    os.remove(output_path)

if __name__ == "__main__":
    test_reproduction()
