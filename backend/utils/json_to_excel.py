# backend/utils/json_to_excel.py

import json
from typing import List, Dict, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def dynamic_json_to_excel(
    json_data: List[Dict], 
    output_path: str, 
    header_map: Optional[Dict[str, str]] = None,
    hidden_columns: Optional[List[str]] = None,
    column_order: Optional[List[str]] = None
) -> str:
    """
    Dynamically converts a list of JSON objects (dictionaries) into a
    formatted Excel file. The columns are inferred from the keys of the
    first JSON object, and can be renamed using the header_map.

    Args:
        json_data: A list of dictionaries, where each dictionary represents a row.
        output_path: The full path where the Excel file will be saved.
        header_map: An optional dictionary to rename columns. 
                    Example: {'guideline_1_summary': 'Old Guideline.xlsx'}
        hidden_columns: An optional list of column keys to exclude from the Excel file.
        column_order: An optional list of column keys to specify the order of columns.

    Returns:
        The path to the created Excel file.
    """
    
    if not json_data:
        print("⚠️ No data to write to Excel. Creating an empty file.")
        wb = Workbook()
        ws = wb.active
        ws['A1'] = "No structured data was extracted or generated."
        wb.save(output_path)
        return output_path

    print(f"📊 Dynamically converting {len(json_data)} items to Excel...")

    # Infer original headers from the keys of the first valid object
    try:
        original_headers_raw = list(json_data[0].keys())
    except (IndexError, AttributeError):
        return dynamic_json_to_excel([], output_path) # Handle empty/invalid data

    hidden_columns = hidden_columns or []

    # ✅ Define preferred column order for ingestion results (fallback if column_order not provided)
    preferred_order = column_order or ["category", "sub_category", "page_number", "guideline_summary"]
    
    # Order columns: preferred fields first, then any additional fields
    original_headers = []
    for field in preferred_order:
        if field in original_headers_raw and field not in hidden_columns:
            original_headers.append(field)
    
    # Add any remaining fields not in the preferred order
    for field in original_headers_raw:
        if field not in original_headers and field not in hidden_columns:
            original_headers.append(field)

    # Create final headers, renaming if a map is provided
    header_map = header_map or {}
    final_headers = []

    for h in original_headers:
        final_headers.append(header_map.get(h, h.replace("_", " ").title()))
    
    print(f"   - Inferred Columns: {original_headers}")
    print(f"   - Final Excel Headers: {final_headers}")

    wb = Workbook()
    ws = wb.active
    ws.title = "Results"

    # --- Styling ---
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border_thin = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    # --- Write Headers ---
    for col_num, header_title in enumerate(final_headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header_title
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border_thin

    # --- Write Data Rows ---
    for row_num, item in enumerate(json_data, 2):
        for col_num, header_key in enumerate(original_headers, 1):
            cell = ws.cell(row=row_num, column=col_num)
            value = item.get(header_key)
            
            # Handle complex data types like lists or dicts
            if isinstance(value, (list, dict)):
                try:
                    cell.value = json.dumps(value, indent=2)
                except TypeError:
                    cell.value = str(value)
            else:
                cell.value = str(value) if value is not None else ""
            
            cell.border = border_thin
            cell.alignment = Alignment(wrap_text=True, vertical='top')

    # --- Auto-fit Column Widths ---
    for col_num, header_key in enumerate(original_headers, 1):
        column_letter = get_column_letter(col_num)
        
        # Calculate maximum character length in this column
        # Start with header length
        max_length = len(str(header_map.get(header_key, header_key.replace("_", " ").title())))
        
        # Check first 100 rows for performance (or all if less)
        sample_rows = json_data[:100]
        for item in sample_rows:
            val = item.get(header_key)
            if val:
                # Handle lists/dicts formatting
                if isinstance(val, (list, dict)):
                    val_str = json.dumps(val)
                else:
                    val_str = str(val)
                
                # Use split lines to find the longest line in case of wrap_text
                longest_line = max([len(line) for line in val_str.split('\n')]) if val_str else 0
                max_length = max(max_length, longest_line)
        
        # Calculate width: max_length + padding, capped at 100 for readability
        adjusted_width = min(100, max_length + 4)
        
        # Override for specific columns
        if header_key == "page_number":
            adjusted_width = 12
        elif header_key == "s_no" or header_key == "sno":
            adjusted_width = 8
            
        ws.column_dimensions[column_letter].width = adjusted_width

    # Freeze the header row
    ws.freeze_panes = 'A2'
    
    # --- Save Workbook ---
    try:
        wb.save(output_path)
        print(f"✅ Dynamic Excel file created successfully: {output_path}")
    except Exception as e:
        print(f"❌ Failed to save Excel file: {e}")
        raise

    return output_path