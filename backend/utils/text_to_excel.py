# utils/text_to_excel.py
import re
from typing import Dict, List, Tuple
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def parse_any_format_to_excel(content: str, output_path: str) -> str:
    """
    Parse markdown table, JSON, or plain text from LLM and convert to Excel.
    
    Args:
        content: Raw text from LLM (table, markdown, JSON, or plain text)
        output_path: Path to save Excel file
    
    Returns:
        Path to created Excel file
    """
    
    print(f"📝 Parsing content ({len(content)} chars)")
    
    # Try to detect and parse format
    table_data = parse_content(content)
    
    print(f"📊 Parsed {len(table_data)} rows")
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Extracted Guidelines"
    
    # Styling
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    section_fill = PatternFill(start_color="E8F4F8", end_color="E8F4F8", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Write headers
    ws['A1'] = "Major Section Title"
    ws['B1'] = "Subsection Title"
    ws['C1'] = "Summary / Key Requirements"
    
    for cell in ['A1', 'B1', 'C1']:
        ws[cell].fill = header_fill
        ws[cell].font = header_font
        ws[cell].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws[cell].border = border
    
    # Write data rows
    row = 2
    for row_data in table_data:
        ws[f'A{row}'] = row_data.get('major_section', '')
        ws[f'B{row}'] = row_data.get('subsection', '')
        ws[f'C{row}'] = row_data.get('summary', '')
        
        # Apply styling
        for col in ['A', 'B', 'C']:
            cell = ws[f'{col}{row}']
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            
            # Highlight major section rows
            if row_data.get('major_section') and not row_data.get('subsection'):
                cell.fill = section_fill
                if col == 'A':
                    cell.font = Font(bold=True, size=11)
        
        row += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 70
    
    # Save workbook
    wb.save(output_path)
    print(f"✅ Excel file created: {output_path}")
    
    return output_path


def parse_content(content: str) -> List[Dict[str, str]]:
    """
    Parse content in any format and return list of row dictionaries.
    
    Returns:
        List of dicts with keys: major_section, subsection, summary
    """
    import json
    
    # Try markdown table format first (most likely from your prompt)
    if '|' in content and ('---' in content or '|---|' in content):
        print("✅ Detected markdown table format")
        return parse_markdown_table(content)
    
    # Try JSON
    try:
        data = json.loads(content.strip())
        if isinstance(data, (dict, list)):
            print("✅ Detected JSON format")
            return parse_json_format(data)
    except:
        pass
    
    # Parse as structured text/markdown
    print("✅ Parsing as structured text format")
    return parse_structured_text(content)


def parse_markdown_table(content: str) -> List[Dict[str, str]]:
    """
    Parse markdown table format.
    
    Example:
    | Major Section | Subsection | Summary |
    |---|---|---|
    | Section 1 | Subsection A | Details here |
    """
    rows = []
    lines = content.split('\n')
    
    in_table = False
    header_found = False
    
    for line in lines:
        line = line.strip()
        
        # Skip empty lines
        if not line:
            continue
        
        # Check if this is a table row
        if line.startswith('|') and line.endswith('|'):
            # Split by |
            cells = [cell.strip() for cell in line.split('|')[1:-1]]
            
            # Skip separator rows (|---|---|---|)
            if all(re.match(r'^-+$', cell.strip()) for cell in cells if cell.strip()):
                in_table = True
                header_found = True
                continue
            
            # Skip header row
            if not header_found:
                continue
            
            # Parse data row
            if in_table and len(cells) >= 3:
                rows.append({
                    'major_section': cells[0].strip(),
                    'subsection': cells[1].strip() if len(cells) > 1 else '',
                    'summary': cells[2].strip() if len(cells) > 2 else ''
                })
    
    print(f"   Parsed {len(rows)} table rows")
    return rows


def parse_json_format(data) -> List[Dict[str, str]]:
    """Parse JSON format"""
    rows = []
    
    if isinstance(data, list):
        for item in data:
            if isinstance(item, dict):
                rows.append({
                    'major_section': item.get('major_section', item.get('section', '')),
                    'subsection': item.get('subsection', ''),
                    'summary': item.get('summary', item.get('details', ''))
                })
    elif isinstance(data, dict):
        for section, content in data.items():
            if isinstance(content, dict):
                # Add section summary if exists
                if 'summary' in content:
                    rows.append({
                        'major_section': section,
                        'subsection': '',
                        'summary': content['summary']
                    })
                
                # Add subsections
                for key, value in content.items():
                    if key != 'summary':
                        rows.append({
                            'major_section': section,
                            'subsection': key,
                            'summary': str(value)
                        })
            else:
                rows.append({
                    'major_section': section,
                    'subsection': '',
                    'summary': str(content)
                })
    
    return rows


def parse_structured_text(content: str) -> List[Dict[str, str]]:
    """Parse structured text/markdown with sections and bullets"""
    rows = []
    current_section = None
    
    lines = content.split('\n')
    
    for line in lines:
        line = line.strip()
        if not line:
            continue
        
        # Detect major sections (## Heading or **Bold**)
        if line.startswith('##'):
            current_section = line.lstrip('#').strip()
            rows.append({
                'major_section': current_section,
                'subsection': '',
                'summary': ''
            })
        elif line.startswith('**') and line.endswith('**'):
            current_section = line.strip('*').strip()
            rows.append({
                'major_section': current_section,
                'subsection': '',
                'summary': ''
            })
        # Detect subsections (bullet points with :)
        elif line.startswith('*') or line.startswith('-'):
            bullet = line.lstrip('*-').strip()
            if ':' in bullet:
                subsection, summary = bullet.split(':', 1)
                rows.append({
                    'major_section': current_section or '',
                    'subsection': subsection.strip(),
                    'summary': summary.strip()
                })
            else:
                rows.append({
                    'major_section': current_section or '',
                    'subsection': '',
                    'summary': bullet
                })
    
    # If no structure found, just add the content
    if not rows:
        rows.append({
            'major_section': 'Extracted Content',
            'subsection': '',
            'summary': content[:1000]  # First 1000 chars
        })
    
    return rows