import asyncio
import datetime
import os
import sys
from typing import List
from dotenv import load_dotenv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Add parent dir to sys.path to allow imports from backend
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from backend.dscr_rules_engine import get_dscr_rules, DSCRRule
from backend.chat.rag_service import RAGService
from backend.utils.llm_provider import LLMProvider
from backend.config import SUPPORTED_MODELS

# Load environment variables
load_dotenv(os.path.join(os.path.dirname(__file__), ".env"))

# Configuration
OUTPUT_FILE = r"C:\Users\LDNA40022\Lokesh\GuidelineIQ\dscr_rag_output.xlsx"
LIMIT_ROWS = None # Set to integer (e.g., 5) for testing, None for full run
MODEL_NAME = "gemini-2.5-flash" # Use fast model for batch processing

async def process_rules():
    print("Starting RAG Excel Generation...")
    
    # 1. Initialize Services
    try:
        rag_service = RAGService()
        
        # Check available keys
        api_key = os.getenv("GEMINI_API_KEY")
        if not api_key:
            print("GEMINI_API_KEY not found in .env")
            return

        llm = LLMProvider(
            provider="gemini",
            api_key=api_key,
            model=MODEL_NAME
        )
        print(f"Services Initialized (Model: {MODEL_NAME})")
    except Exception as e:
        print(f"Failed to initialize services: {e}")
        return

    # 2. Get Rules
    rules = get_dscr_rules()
    if LIMIT_ROWS:
        rules = rules[:LIMIT_ROWS]
        print(f"Limit applied: Processing only {LIMIT_ROWS} rules.")
    
    print(f"Processing {len(rules)} rules...")

    # 3. RAG Loop
    results = []
    
    # Semaphore for concurrency
    semaphore = asyncio.Semaphore(5)

    async def fetch_rule_content(rule: DSCRRule):
        async with semaphore:
            query = f"What are the requirements for {rule.sub_categories[0]} regarding {rule.dscr_parameter}? {rule.variance_category}"
            
            try:
                # Search Vector DB
                search_results = await rag_service.search(
                    query=query,
                    provider="gemini",
                    api_key=api_key,
                    n_results=5
                )
                
                context_text = ""
                if search_results:
                    context_text = "\n\n".join([r['text'] for r in search_results])
                
                # If no context, return empty or default note
                if not context_text:
                    return {
                        "rule": rule,
                        "content": "No specific context found in knowledge base."
                    }

                # Summarize with LLM
                prompt = f"""
                You are a Mortgage Policy Summarizer.
                Initial Request: {query}
                
                Context from Guidelines:
                {context_text}
                
                Task:
                1. Extract the specific "Variance Category" and "SubCategory" that this guideline belongs to based strictly on the context.
                2. Create a bulleted list of the specific requirements, limits, and conditions for "{rule.dscr_parameter}" based strictly on the context.
                
                Format your response as a JSON object with the following keys:
                - "variance_category": (string)
                - "subcategory": (string)
                - "summary": (string, clean list with "• " bullets)

                Be concise. If the context doesn't explicitly mention something, state "NA".
                """
                
                response_text = await asyncio.to_thread(
                    llm.generate,
                    "You are a helpful mortgage expert assistant. Always return valid JSON.",
                    prompt
                )
                
                import json
                try:
                    # Basic JSON cleaning in case LLM adds markdown blocks
                    json_str = response_text.strip()
                    if json_str.startswith("```json"):
                        json_str = json_str[7:]
                    if json_str.endswith("```"):
                        json_str = json_str[:-3]
                    
                    data_json = json.loads(json_str.strip())
                    return {
                        "rule": rule,
                        "rag_variance": data_json.get("variance_category", "NA"),
                        "rag_subcat": data_json.get("subcategory", "NA"),
                        "content": data_json.get("summary", "No summary provided.")
                    }
                except Exception as json_err:
                    print(f"JSON Parse Error for {rule.dscr_parameter}: {json_err}")
                    return {
                        "rule": rule,
                        "rag_variance": "Error parsing",
                        "rag_subcat": "Error parsing",
                        "content": response_text.strip()
                    }

            except Exception as e:
                print(f"Error processing {rule.dscr_parameter}: {e}")
                return {
                    "rule": rule,
                    "rag_variance": "Error",
                    "rag_subcat": "Error",
                    "content": "Error retrieval content."
                }

    # Run tasks
    tasks = [fetch_rule_content(r) for r in rules]
    processed_data = []
    
    # Progress indicator
    total = len(tasks)
    for i, fut in enumerate(asyncio.as_completed(tasks)):
        result = await fut
        processed_data.append(result)
        print(f"[{i+1}/{total}] Processed: {result['rule'].dscr_parameter}")

    # 4. Generate Excel
    create_excel_file(processed_data)

def create_excel_file(data: List[dict]):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DSCR 1-4 RAG Output"

    # Exact 5 headers matching the user's image
    headers = [
        "DSCR Parameters\n(Investor / Business Purpose Loans)", 
        "Variance Categories", 
        "SubCategories", 
        "PPE Field Type", 
        f"NQMF Investor DSCR (1-4 Units) Generated {datetime.date.today()}"
    ]
    
    ws.append(headers)

    # Styles
    header_font = Font(bold=True, size=11)
    header_fill_blue = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid") # Light Blue
    header_fill_orange = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid") # Peach/Orange (Column B, C, D)
    
    # Apply Header Styles
    for col_num, cell in enumerate(ws[1], 1):
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        # Apply colors based on column index
        if col_num == 1:
            cell.fill = header_fill_blue
        elif col_num in [2, 3, 4]:
            cell.fill = header_fill_orange
        else: # Content column (Column E)
            cell.fill = header_fill_blue

    # Restore Order
    original_order = {r.dscr_parameter: i for i, r in enumerate(get_dscr_rules())}
    data.sort(key=lambda x: original_order.get(x['rule'].dscr_parameter, 999))

    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
    
    fill_light_yellow = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    fill_light_green = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    for item in data:
        rule = item['rule']
        content = item['content']
        # Use RAG extracted fields for the main category columns
        rag_variance = item.get('rag_variance', 'Not extracted')
        rag_subcat = item.get('rag_subcat', 'Not extracted')
        
        row = [
            rule.dscr_parameter, # Column A
            rag_variance,        # Column B (Now RAG)
            rag_subcat,          # Column C (Now RAG)
            rule.policy_type,    # Column D
            content              # Column E
        ]
        ws.append(row)
        
        # Apply row styles
        current_row = ws.max_row
        for col_idx, cell in enumerate(ws[current_row], 1):
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = thin_border
            
            # Fill colors based on column
            if col_idx in [1, 2, 3]:
                cell.fill = fill_light_green # Green for Parameters and Categories
            elif col_idx == 4:
                # Keep Field Type blank or neutral
                pass
            elif col_idx == 5:
                cell.fill = fill_light_yellow # Yellow for Content/Matrix

    # Column Widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 80

    try:
        wb.save(OUTPUT_FILE)
        print(f"Excel generated successfully: {OUTPUT_FILE}")
    except Exception as e:
        print(f"Error saving Excel: {e}")

if __name__ == "__main__":
    if len(sys.argv) > 1 and sys.argv[1] == "--limit":
         LIMIT_ROWS = int(sys.argv[2])
    
    asyncio.run(process_rules())
