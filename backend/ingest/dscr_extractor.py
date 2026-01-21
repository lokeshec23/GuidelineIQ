# backend/ingest/dscr_extractor.py

import os
import asyncio
import json
import datetime
from typing import List, Dict, Tuple
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from ingest.dscr_config import DSCR_GUIDELINES
from chat.rag_service import RAGService
from utils.llm_provider import LLMProvider

async def extract_dscr_parameters_safe(
    session_id: str,
    gridfs_file_id: str,
    rag_service: RAGService,
    llm: LLMProvider,
    investor: str,
    version: str,
    user_settings: dict
) -> Tuple[str, List[Dict]]:
    """
    Extracts DSCR parameters using RAG and LLM, and saves them to an Excel file.
    Returns:
        Tuple[str, List[Dict]]: (The path to the generated Excel file, The data list)
    """
    print(f"\n{'='*60}")
    print(f"🚀 Starting RAG-Based DSCR Extraction for Session: {session_id[:8]}")
    print(f"{'='*60}\n")
    
    # Concurrency control
    semaphore = asyncio.Semaphore(10)
    
    async def process_one(guideline_config: Dict) -> Dict:
        async with semaphore:
            param = guideline_config["parameter"]
            # Hardcoded values
            category = guideline_config.get("category", "General")
            subcategory = guideline_config.get("subcategory", "General")
            ppe_field = guideline_config.get("ppe_field", "Text")
            
            try:
                # Setup provider/key (reusing logic from thought process)
                provider = llm.provider 
                api_key = llm.api_key # The LLMProvider has the key stored
                
                filter_metadata = {"gridfs_file_id": gridfs_file_id}
                
                # Check for aliases
                base_query = f"What are the requirements for {param}?"
                search_query = base_query
                
                aliases = guideline_config.get("aliases", [])
                if aliases:
                    search_query = f"{param} {' '.join(aliases)}"

                search_results = await rag_service.search(
                    query=search_query,
                    provider=provider,
                    api_key=api_key,
                    n_results=5,
                    filter_metadata=filter_metadata,
                    azure_endpoint=user_settings.get("openai_endpoint"),
                    azure_embedding_deployment=user_settings.get("openai_embedding_deployment", "embedding-model")
                )
                
                context_text = ""
                if search_results:
                    # ✅ Enhanced: Include filename and page number in source attribution
                    context_text = "\n\n".join([
                        f"[Source: {r['metadata'].get('filename', 'Unknown')} - Page {r['metadata'].get('page', '?')}]\n{r['text']}" 
                        for r in search_results
                    ])
                
                if not context_text:
                    return {
                        "DSCR_Parameters": param, 
                        "Variance_Category": category,
                        "SubCategory": subcategory,
                        "PPE_Field_Type": ppe_field,
                        "NQMF Investor DSCR": "NA"
                    }
                
                # Enhanced Prompt for Detailed Extraction
                # Note: We no longer ask for category/subcategory since they are hardcoded
                system_prompt = "You are a Mortgage Policy Summarizer."
                user_msg = f"""
                Initial Request: {base_query}
                
                Context from Guidelines:
                {context_text}
                
                Task:
                Create a bulleted list of the specific requirements, limits, and conditions for "{param}" based strictly on the context.
                
                Format your response as a JSON object with the following key:
                - "summary": (string, clean list with "• " bullets)

                Be concise. If the context doesn't explicitly mention something, state "NA".
                """
                
                response_text = await asyncio.to_thread(
                    llm.generate,
                    "You are a helpful mortgage expert assistant. Always return valid JSON.",
                    user_msg
                )
                
                try:
                    # Basic JSON cleaning
                    json_str = response_text.strip()
                    if json_str.startswith("```json"):
                        json_str = json_str[7:]
                    if json_str.endswith("```"):
                        json_str = json_str[:-3]
                    
                    data_json = json.loads(json_str.strip())
                    return {
                        "DSCR_Parameters": param,
                        "Variance_Category": category,
                        "SubCategory": subcategory,
                        "PPE_Field_Type": ppe_field,
                        "NQMF Investor DSCR": data_json.get("summary", "No summary provided.")
                    }
                except Exception as json_err:
                    print(f"JSON Parse Error for {param}: {json_err}")
                    return {
                        "DSCR_Parameters": param,
                        "Variance_Category": category,
                        "SubCategory": subcategory,
                        "PPE_Field_Type": ppe_field,
                        "NQMF Investor DSCR": response_text.strip()
                    }
                
            except Exception as e:
                print(f"Error on {param}: {e}")
                return {
                    "DSCR_Parameters": param,
                    "Variance_Category": category,
                    "SubCategory": subcategory, 
                    "PPE_Field_Type": ppe_field,
                    "NQMF Investor DSCR": "Error extraction"
                }

    # Execute
    tasks = [process_one(g) for g in DSCR_GUIDELINES]
    results = await asyncio.gather(*tasks)
    
    # Generate Excel
    filepath = create_dscr_excel(results, session_id, investor, version)
    return filepath, results


async def extract_dscr_parameters_multi_pdf(
    session_id: str,
    gridfs_file_ids: List[str],
    filenames: List[str],
    rag_service: RAGService,
    llm: LLMProvider,
    investor: str,
    version: str,
    user_settings: dict
) -> Tuple[str, List[Dict]]:
    """
    Extracts DSCR parameters from multiple PDFs and aggregates results.
    
    Args:
        session_id: Unique session identifier
        gridfs_file_ids: List of GridFS file IDs for the PDFs
        filenames: List of original PDF filenames
        rag_service: RAG service instance
        llm: LLM provider instance
        investor: Investor name
        version: Version identifier
        user_settings: User configuration settings
    
    Returns:
        Tuple[str, List[Dict]]: (Excel file path, Aggregated results list)
    """
    print(f"\n{'='*60}")
    print(f"🚀 Starting Multi-PDF DSCR Extraction for Session: {session_id[:8]}")
    print(f"Processing {len(gridfs_file_ids)} PDFs")
    print(f"{'='*60}\n")
    
    # Store all results by parameter name
    aggregated_by_parameter = {}
    
    # ✅ PARALLEL PROCESSING: Process all PDFs concurrently
    async def process_single_pdf(idx: int, file_id: str, filename: str):
        """Process a single PDF and return its results"""
        print(f"\n📄 Processing PDF {idx}/{len(gridfs_file_ids)}: {filename}")
        print(f"{'='*60}")
        
        try:
            # Extract DSCR parameters for this PDF
            _, pdf_results = await extract_dscr_parameters_safe(
                session_id=session_id,
                gridfs_file_id=file_id,
                rag_service=rag_service,
                llm=llm,
                investor=investor,
                version=version,
                user_settings=user_settings
            )
            
            print(f"✅ Completed extraction for {filename}")
            return (filename, pdf_results)
            
        except Exception as e:
            print(f"❌ Error processing {filename}: {e}")
            return (filename, [])
    
    # Execute all PDF extractions in parallel
    tasks = [
        process_single_pdf(idx, file_id, filename)
        for idx, (file_id, filename) in enumerate(zip(gridfs_file_ids, filenames), 1)
    ]
    
    all_results = await asyncio.gather(*tasks)
    
    # Aggregate results from all PDFs
    for filename, pdf_results in all_results:
        if not pdf_results:
            continue
            
        for result in pdf_results:
            param_name = result["DSCR_Parameters"]
            
            if param_name not in aggregated_by_parameter:
                aggregated_by_parameter[param_name] = {
                    "parameter": param_name,
                    "category": result.get("Variance_Category", ""),
                    "subcategory": result.get("SubCategory", ""),
                    "ppe_field": result.get("PPE_Field_Type", ""),
                    "extractions": []
                }
            
            # Store extraction with source info
            aggregated_by_parameter[param_name]["extractions"].append({
                "source_pdf": filename,
                "summary": result["NQMF Investor DSCR"]
            })
    
    # Summarize aggregated results
    print(f"\n{'='*60}")
    print(f"🔄 Summarizing results from {len(gridfs_file_ids)} PDFs...")
    print(f"{'='*60}\n")
    
    final_results = await summarize_dscr_aggregated_results(
        aggregated_by_parameter=aggregated_by_parameter,
        llm=llm
    )
    
    # Generate Excel with multi-PDF context
    filepath = create_dscr_excel_multi_pdf(
        data=final_results,
        session_id=session_id,
        investor=investor,
        version=version,
        pdf_filenames=filenames
    )
    
    return filepath, final_results


async def summarize_dscr_aggregated_results(
    aggregated_by_parameter: Dict[str, Dict],
    llm: LLMProvider
) -> List[Dict]:
    """
    Uses LLM to intelligently summarize DSCR parameters extracted from multiple PDFs.
    
    Args:
        aggregated_by_parameter: Dictionary mapping parameter names to their extractions
        llm: LLM provider for summarization
    
    Returns:
        List of dictionaries with summarized DSCR parameters
    """
    final_results = []
    semaphore = asyncio.Semaphore(5)  # Limit concurrent summarizations
    
    async def summarize_one(param_name: str, param_data: Dict) -> Dict:
        async with semaphore:
            extractions = param_data["extractions"]
            
            # If only one extraction or all are "NA", no summarization needed
            unique_summaries = set(e["summary"] for e in extractions)
            
            if len(extractions) == 1:
                return {
                    "DSCR_Parameters": param_name,
                    "Variance_Category": param_data["category"],
                    "SubCategory": param_data["subcategory"],
                    "PPE_Field_Type": param_data["ppe_field"],
                    "NQMF Investor DSCR": extractions[0]["summary"]
                }
            
            # Check if all are "NA" or similar
            if all("na" == s.lower().strip() or "not found" in s.lower() or "error" in s.lower() for s in unique_summaries):
                return {
                    "DSCR_Parameters": param_name,
                    "Variance_Category": param_data["category"],
                    "SubCategory": param_data["subcategory"],
                    "PPE_Field_Type": param_data["ppe_field"],
                    "NQMF Investor DSCR": "NA"
                }
            
            # Build context for LLM
            context_parts = []
            for i, extraction in enumerate(extractions, 1):
                context_parts.append(f"**PDF {i}: {extraction['source_pdf']}**\n{extraction['summary']}")
            
            context_text = "\n\n".join(context_parts)
            
            # Prompt for intelligent summarization
            system_prompt = "You are an expert at analyzing and summarizing mortgage lending guidelines."
            user_prompt = f"""
You have extracted information about the DSCR parameter "{param_name}" from {len(extractions)} different PDF documents.

Below are the extractions from each PDF:

{context_text}

Task:
Create a comprehensive, unified summary that:
1. Combines all relevant information from the PDFs
2. Highlights any differences or conflicts between documents
3. Uses clear bullet points (• ) for readability
4. Prioritizes the most restrictive or specific requirements when there are conflicts
5. Indicates the source PDF(s) for critical requirements using (PDF 1), (PDF 2), etc.

Format your response as a JSON object with this key:
- "summary": (string, the unified summary with bullet points)

Be concise but complete. If information is consistent across PDFs, state it once.
"""
            
            try:
                response_text = await asyncio.to_thread(
                    llm.generate,
                    system_prompt,
                    user_prompt
                )
                
                # Parse JSON response
                json_str = response_text.strip()
                if json_str.startswith("```json"):
                    json_str = json_str[7:]
                if json_str.endswith("```"):
                    json_str = json_str[:-3]
                
                data_json = json.loads(json_str.strip())
                summarized_text = data_json.get("summary", response_text.strip())
                
            except Exception as e:
                print(f"⚠️ Summarization failed for {param_name}: {e}")
                # Fallback: concatenate all summaries
                summarized_text = "\n\n".join([
                    f"**From {e['source_pdf']}:**\n{e['summary']}" 
                    for e in extractions
                ])
            
            return {
                "DSCR_Parameters": param_name,
                "Variance_Category": param_data["category"],
                "SubCategory": param_data["subcategory"],
                "PPE_Field_Type": param_data["ppe_field"],
                "NQMF Investor DSCR": summarized_text
            }
    
    # Summarize all parameters concurrently
    tasks = [
        summarize_one(param_name, param_data) 
        for param_name, param_data in aggregated_by_parameter.items()
    ]
    final_results = await asyncio.gather(*tasks)
    
    # Sort by parameter order from config
    param_order = {item['parameter']: i for i, item in enumerate(DSCR_GUIDELINES)}
    final_results.sort(key=lambda x: param_order.get(x['DSCR_Parameters'], 999))
    
    print(f"✅ Summarization complete for {len(final_results)} parameters")
    return final_results

def create_dscr_excel(data: List[Dict], session_id: str, investor: str, version: str) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "DSCR 1-4 RAG Output"

    # Exact 5 headers 
    headers = [
        "DSCR Parameters\n(Investor / Business Purpose Loans)", 
        "Variance Categories", 
        "SubCategories", 
        "PPE Field Type", 
        f"NQMF Investor DSCR (1-4 Units) Generated {datetime.date.today()}"
    ]
    
    ws.append(headers)

    # Styles
    header_font = Font(bold=True, size=11)
    header_fill_blue = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid") # Light Blue
    header_fill_orange = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid") # Peach/Orange
    
    # Apply Header Styles
    for col_num, cell in enumerate(ws[1], 1):
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        # Apply colors based on column index
        if col_num == 1:
            cell.fill = header_fill_blue
        elif col_num in [2, 3, 4]:
            cell.fill = header_fill_orange
        else: # Content column
            cell.fill = header_fill_blue

    # Sort results to match config order
    # Create a map of param -> index from config
    param_order = {item['parameter']: i for i, item in enumerate(DSCR_GUIDELINES)}
    data.sort(key=lambda x: param_order.get(x['DSCR_Parameters'], 999))

    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
    
    fill_light_yellow = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    fill_light_green = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    for item in data:
        row = [
            item['DSCR_Parameters'],       # Column A
            item.get('Variance_Category', ''),  # Column B
            item.get('SubCategory', ''),    # Column C
            item.get('PPE_Field_Type', ''), # Column D
            item['NQMF Investor DSCR']     # Column E
        ]
        ws.append(row)
        
        # Apply row styles
        current_row = ws.max_row
        for col_idx, cell in enumerate(ws[current_row], 1):
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = thin_border
            
            # Fill colors based on column
            if col_idx in [1, 2, 3]:
                cell.fill = fill_light_green 
            elif col_idx == 4:
                pass
            elif col_idx == 5:
                cell.fill = fill_light_yellow 

    # Column Widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 80
    
    # File Path
    filename = f"DSCR_Extraction_{investor}_{version}_{session_id[:8]}.xlsx"
    
    # Save to 'results' folder in backend
    base_dir = os.getcwd()
    results_dir = os.path.join(base_dir, "results")
    if not os.path.exists(results_dir):
        os.makedirs(results_dir)
        
    filepath = os.path.join(results_dir, filename)
    
    try:
        wb.save(filepath)
        print(f"✅ DSCR Excel saved at: {filepath}")
        return filepath
    except Exception as e:
        print(f"Error saving Excel: {e}")
        # Fallback to temp if results dir fails
        return f"Error: {e}"

def create_dscr_excel_multi_pdf(
    data: List[Dict], 
    session_id: str, 
    investor: str, 
    version: str,
    pdf_filenames: List[str]
) -> str:
    """
    Creates an Excel file for multi-PDF DSCR extraction with metadata section.
    
    Args:
        data: List of DSCR parameter dictionaries with summarized values
        session_id: Session identifier
        investor: Investor name
        version: Version identifier
        pdf_filenames: List of source PDF filenames
    
    Returns:
        Path to the generated Excel file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "DSCR Multi-PDF Output"
    
    # === METADATA SECTION ===
    # Title row
    ws.append([f"Multi-PDF DSCR Extraction Report - {len(pdf_filenames)} Documents"])
    ws.merge_cells('A1:E1')
    title_cell = ws['A1']
    title_cell.font = Font(bold=True, size=14, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Metadata info
    ws.append([f"Investor: {investor}"])
    ws.append([f"Version: {version}"])
    ws.append([f"Generated: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
    ws.append([f"Number of PDFs Processed: {len(pdf_filenames)}"])
    
    # Source PDFs list
    ws.append(["Source PDF Files:"])
    ws['A6'].font = Font(bold=True)
    for idx, filename in enumerate(pdf_filenames, 1):
        ws.append([f"  {idx}. {filename}"])
    
    # Empty row separator
    ws.append([])
    
    # === DATA SECTION ===
    headers_row = ws.max_row + 1
    headers = [
        "DSCR Parameters\n(Investor / Business Purpose Loans)", 
        "Variance Categories", 
        "SubCategories", 
        "PPE Field Type", 
        f"NQMF Investor DSCR (Aggregated from {len(pdf_filenames)} PDFs)"
    ]
    
    ws.append(headers)

    # Styles
    header_font = Font(bold=True, size=11)
    header_fill_blue = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    header_fill_orange = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")
    
    # Apply Header Styles
    for col_num, cell in enumerate(ws[headers_row], 1):
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        if col_num == 1:
            cell.fill = header_fill_blue
        elif col_num in [2, 3, 4]:
            cell.fill = header_fill_orange
        else:  # Content column
            cell.fill = header_fill_blue

    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
    
    fill_light_yellow = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    fill_light_green = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    for item in data:
        row = [
            item['DSCR_Parameters'],
            item.get('Variance_Category', ''),
            item.get('SubCategory', ''),
            item.get('PPE_Field_Type', ''),
            item['NQMF Investor DSCR']
        ]
        ws.append(row)
        
        # Apply row styles
        current_row = ws.max_row
        for col_idx, cell in enumerate(ws[current_row], 1):
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = thin_border
            
            # Fill colors based on column
            if col_idx in [1, 2, 3]:
                cell.fill = fill_light_green 
            elif col_idx == 4:
                pass
            elif col_idx == 5:
                cell.fill = fill_light_yellow 

    # Column Widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 80
    
    # File Path
    filename = f"DSCR_MultiPDF_{investor}_{version}_{session_id[:8]}.xlsx"
    
    # Save to 'results' folder in backend
    base_dir = os.getcwd()
    results_dir = os.path.join(base_dir, "results")
    if not os.path.exists(results_dir):
        os.makedirs(results_dir)
        
    filepath = os.path.join(results_dir, filename)
    
    try:
        wb.save(filepath)
        print(f"✅ Multi-PDF DSCR Excel saved at: {filepath}")
        return filepath
    except Exception as e:
        print(f"Error saving Excel: {e}")
        return f"Error: {e}"

