
import openpyxl
import os

EXCEL_FILE_PATH = os.path.join(os.getcwd(), 'backend', 'guideline_type_parameters.xlsx')

def check_counts():
    wb = openpyxl.load_workbook(EXCEL_FILE_PATH, data_only=True)
    ws = wb.active
    
    total_rows = ws.max_row
    print(f"Total rows in Excel (max_row): {total_rows}")
    
    fa_count = 0
    dscr_count = 0
    other_count = 0
    empty_rows = []
    
    # Check rows from 2 onwards (skipping header)
    for r in range(2, total_rows + 1):
        # Column 4 is "Guideline Type"
        gtype = ws.cell(row=r, column=4).value
        # Column 1 is "Parameters"
        param = ws.cell(row=r, column=1).value
        
        if param is None and gtype is None:
            empty_rows.append(r)
            continue
            
        if gtype == 'Full & Alt':
            fa_count += 1
        elif gtype == 'DSCR':
            dscr_count += 1
        else:
            other_count += 1
            print(f"Row {r}: Other type: {repr(gtype)} (Param: {repr(param)})")
            
    print(f"Full & Alt count: {fa_count}")
    print(f"DSCR count: {dscr_count}")
    print(f"Other/Invalid count: {other_count}")
    print(f"Empty rows detected: {empty_rows}")

if __name__ == "__main__":
    check_counts()
